#!/usr/bin/env python3
"""
Unified export script for Bhulekh RoR data.

Supports CSV and Excel formats, per-district or per-village output.

Usage Examples:
    # Export all districts to CSV (one file per district)
    uv run python export_data.py --data-dir bhulekh_data --output-dir export

    # Export specific districts to CSV
    uv run python export_data.py --data-dir bhulekh_data --output-dir export --districts କଟକ ଅନୁଗୋଳ

    # Export to Excel format
    uv run python export_data.py --data-dir bhulekh_data --output-dir export --format xlsx

    # Export individual village files (organized by district/tahasil/village)
    uv run python export_data.py --data-dir bhulekh_data --output-dir export --by-village

    # Export specific district with village-level files in Excel
    uv run python export_data.py --data-dir bhulekh_data --output-dir export --districts କଟକ --by-village --format xlsx

    # Single combined file for all data
    uv run python export_data.py --data-dir bhulekh_data --output-dir export --single-file
"""

import argparse
import csv
import json
import re
import sqlite3
from pathlib import Path
from typing import Any, Dict, List, Optional, Set

# Regex to remove illegal characters
ILLEGAL_CHARS_RE = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]')

CSV_COLUMNS = [
    'district', 'mouja', 'tehsil', 'thana', 'tehsil_no', 'thana_no',
    'landlord_name', 'khatiyan_sl_no', 'tenant_name', 'status',
    'water_tax', 'tax', 'ses', 'other_ses', 'total',
    'description', 'special_case', 'last_publish_date', 'tax_date',
    'form_no', 'parichheda',
    'plot_plot_no', 'plot_chaka', 'plot_land_type', 'plot_kisam',
    'plot_n_occu', 'plot_e_occu', 'plot_s_occu', 'plot_w_occu',
    'plot_acre', 'plot_decimil', 'plot_hector', 'plot_remarks',
]


def clean_value(value: Any) -> str:
    """Clean a value for export."""
    if value is None:
        return ''
    if isinstance(value, str):
        value = ILLEGAL_CHARS_RE.sub('', value)
        value = value.replace('\x00', '')
        value = value.strip()
        if len(value) > 32000:
            value = value[:32000] + '...'
    return str(value)


def sanitize_filename(name: str) -> str:
    """Make a string safe for use as filename."""
    invalid_chars = '<>:"/\\|?*'
    for char in invalid_chars:
        name = name.replace(char, '_')
    return name.strip() or "unknown"


def get_district_name_from_db(db_path: Path) -> str:
    """Extract district name from a district database."""
    try:
        conn = sqlite3.connect(str(db_path))
        row = conn.execute("SELECT DISTINCT district FROM khatiyans LIMIT 1").fetchone()
        conn.close()
        return row[0] if row else db_path.stem
    except:
        return db_path.stem


def read_sqlite_data(db_path: Path) -> List[Dict[str, Any]]:
    """Read all khatiyan data from a SQLite database."""
    records = []
    try:
        conn = sqlite3.connect(str(db_path))
        cursor = conn.execute(
            "SELECT district, tahasil, village, khatiyan_value, khatiyan_text, data_json FROM khatiyans"
        )
        for row in cursor:
            district, tahasil, village, kh_value, kh_text, data_json = row
            try:
                data = json.loads(data_json)
                data['_district'] = district
                data['_tahasil'] = tahasil
                data['_village'] = village
                data['_khatiyan_value'] = kh_value
                data['_khatiyan_text'] = kh_text
                records.append(data)
            except json.JSONDecodeError:
                pass
        conn.close()
    except Exception as e:
        print(f"Error reading {db_path}: {e}")
    return records


def flatten_khatiyan(kh: Dict[str, Any]) -> List[Dict[str, str]]:
    """Flatten a single khatiyan to rows (one per plot)."""
    plots = kh.get('plots', [])

    base_row = {
        'district': clean_value(kh.get('_district', '') or kh.get('district', '')),
        'mouja': clean_value(kh.get('mouja', '') or kh.get('_village', '')),
        'tehsil': clean_value(kh.get('_tahasil', '') or kh.get('tehsil', '')),
        'thana': clean_value(kh.get('thana', '')),
        'tehsil_no': clean_value(kh.get('tehsil_no', '')),
        'thana_no': clean_value(kh.get('thana_no', '')),
        'landlord_name': clean_value(kh.get('landlord_name', '')),
        'khatiyan_sl_no': clean_value(kh.get('khatiyan_sl_no', '') or kh.get('_khatiyan_text', '')),
        'tenant_name': clean_value(kh.get('tenant_name', '')),
        'status': clean_value(kh.get('status', '')),
        'water_tax': clean_value(kh.get('water_tax', '')),
        'tax': clean_value(kh.get('tax', '')),
        'ses': clean_value(kh.get('ses', '')),
        'other_ses': clean_value(kh.get('other_ses', '')),
        'total': clean_value(kh.get('total', '')),
        'description': clean_value(kh.get('description', '')),
        'special_case': clean_value(kh.get('special_case', '')),
        'last_publish_date': clean_value(kh.get('last_publish_date', '')),
        'tax_date': clean_value(kh.get('tax_date', '')),
        'form_no': clean_value(kh.get('form_no', '')),
        'parichheda': clean_value(kh.get('parichheda', '')),
    }

    empty_plot = {
        'plot_plot_no': '', 'plot_chaka': '', 'plot_land_type': '', 'plot_kisam': '',
        'plot_n_occu': '', 'plot_e_occu': '', 'plot_s_occu': '', 'plot_w_occu': '',
        'plot_acre': '', 'plot_decimil': '', 'plot_hector': '', 'plot_remarks': '',
    }

    rows = []
    if plots:
        for plot in plots:
            row = base_row.copy()
            row.update({
                'plot_plot_no': clean_value(plot.get('plot_no', '')),
                'plot_chaka': clean_value(plot.get('chaka', '')),
                'plot_land_type': clean_value(plot.get('land_type', '')),
                'plot_kisam': clean_value(plot.get('kisam', '')),
                'plot_n_occu': clean_value(plot.get('n_occu', '')),
                'plot_e_occu': clean_value(plot.get('e_occu', '')),
                'plot_s_occu': clean_value(plot.get('s_occu', '')),
                'plot_w_occu': clean_value(plot.get('w_occu', '')),
                'plot_acre': clean_value(plot.get('acre', '')),
                'plot_decimil': clean_value(plot.get('decimil', '')),
                'plot_hector': clean_value(plot.get('hector', '')),
                'plot_remarks': clean_value(plot.get('remarks', '')),
            })
            rows.append(row)
    else:
        row = base_row.copy()
        row.update(empty_plot)
        rows.append(row)

    return rows


def write_csv(rows: List[Dict[str, str]], output_path: Path):
    """Write rows to CSV file."""
    with open(output_path, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.DictWriter(f, fieldnames=CSV_COLUMNS, extrasaction='ignore')
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(rows: List[Dict[str, str]], output_path: Path):
    """Write rows to Excel file."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("ERROR: openpyxl not installed. Run: pip install openpyxl")
        print("Falling back to CSV...")
        csv_path = output_path.with_suffix('.csv')
        write_csv(rows, csv_path)
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    # Write header
    ws.append(CSV_COLUMNS)

    # Style header
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col in range(1, len(CSV_COLUMNS) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Write data rows
    for row in rows:
        ws.append([row.get(col, '') for col in CSV_COLUMNS])

    # Auto-width columns
    for col in range(1, len(CSV_COLUMNS) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14

    ws.freeze_panes = 'A2'
    wb.save(output_path)


def write_file(rows: List[Dict[str, str]], output_path: Path, fmt: str):
    """Write rows to file in specified format."""
    if fmt == 'xlsx':
        write_xlsx(rows, output_path)
    else:
        write_csv(rows, output_path)


def export_data(
    data_dir: str,
    output_dir: str,
    fmt: str = 'csv',
    districts: Optional[List[str]] = None,
    by_village: bool = False,
    single_file: bool = False,
):
    """
    Export data with flexible options.
    
    Args:
        data_dir: Directory containing district SQLite files
        output_dir: Output directory
        fmt: Output format ('csv' or 'xlsx')
        districts: Optional list of district names to filter
        by_village: If True, create one file per village
        single_file: If True, create single combined file
    """
    data_path = Path(data_dir)
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    if not data_path.exists():
        print(f"ERROR: Data directory not found: {data_dir}")
        return

    db_files = sorted(data_path.glob("district_*.db"))
    if not db_files:
        print(f"No district databases found in {data_dir}")
        return

    print(f"Found {len(db_files)} district databases")
    if districts:
        print(f"Filtering to districts: {', '.join(districts)}")

    ext = '.' + fmt
    all_rows = []
    total_files = 0
    total_rows = 0

    for db_path in db_files:
        district_name = get_district_name_from_db(db_path)

        if districts and district_name not in districts:
            continue

        print(f"\nProcessing {district_name}...")
        records = read_sqlite_data(db_path)
        print(f"  Loaded {len(records)} khatiyans")

        if by_village:
            # Group by tahasil/village
            by_location = {}
            for kh in records:
                tahasil = kh.get('_tahasil', 'Unknown')
                village = kh.get('_village', 'Unknown')
                key = (tahasil, village)
                if key not in by_location:
                    by_location[key] = []
                by_location[key].append(kh)

            # Create directory structure: district/tahasil/village.ext
            district_dir = output_path / sanitize_filename(district_name)
            
            for (tahasil, village), khatiyans in by_location.items():
                tahasil_dir = district_dir / sanitize_filename(tahasil)
                tahasil_dir.mkdir(parents=True, exist_ok=True)

                rows = []
                for kh in khatiyans:
                    rows.extend(flatten_khatiyan(kh))

                file_path = tahasil_dir / (sanitize_filename(village) + ext)
                write_file(rows, file_path, fmt)
                total_files += 1
                total_rows += len(rows)

            print(f"  Created {len(by_location)} village files")

        elif single_file:
            # Collect all rows for combined file
            for kh in records:
                all_rows.extend(flatten_khatiyan(kh))

        else:
            # One file per district
            rows = []
            for kh in records:
                rows.extend(flatten_khatiyan(kh))

            file_path = output_path / (sanitize_filename(district_name) + ext)
            write_file(rows, file_path, fmt)
            total_files += 1
            total_rows += len(rows)
            print(f"  Wrote {file_path.name} ({len(rows)} rows)")

    if single_file and all_rows:
        file_path = output_path / ('all_districts' + ext)
        write_file(all_rows, file_path, fmt)
        total_files = 1
        total_rows = len(all_rows)
        print(f"\nWrote combined file: {file_path} ({len(all_rows)} rows)")

    print(f"\n{'='*60}")
    print("EXPORT COMPLETE")
    print(f"  Format: {fmt.upper()}")
    print(f"  Files created: {total_files}")
    print(f"  Total rows: {total_rows:,}")
    print(f"  Output: {output_path}")


def main():
    parser = argparse.ArgumentParser(
        description="Export Bhulekh RoR data to CSV or Excel",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Export all to CSV (one file per district)
  uv run python export_data.py --data-dir bhulekh_data --output-dir export

  # Export specific districts
  uv run python export_data.py --data-dir bhulekh_data --districts କଟକ ଅନୁଗୋଳ

  # Export to Excel
  uv run python export_data.py --data-dir bhulekh_data --format xlsx

  # Export individual village files
  uv run python export_data.py --data-dir bhulekh_data --by-village

  # Combined single file
  uv run python export_data.py --data-dir bhulekh_data --single-file
"""
    )
    parser.add_argument(
        "--data-dir", default="bhulekh_data",
        help="Directory containing district SQLite files (default: bhulekh_data)"
    )
    parser.add_argument(
        "--output-dir", default="export",
        help="Output directory (default: export)"
    )
    parser.add_argument(
        "--format", "-f", choices=['csv', 'xlsx'], default='csv',
        help="Output format: csv or xlsx (default: csv)"
    )
    parser.add_argument(
        "--districts", nargs='+', default=None,
        help="Filter to specific district names (Odia text, space-separated)"
    )
    parser.add_argument(
        "--by-village", action="store_true",
        help="Create individual files per village (organized in folders)"
    )
    parser.add_argument(
        "--single-file", action="store_true",
        help="Export all data to a single combined file"
    )

    args = parser.parse_args()

    if args.by_village and args.single_file:
        print("ERROR: Cannot use --by-village and --single-file together")
        return

    export_data(
        data_dir=args.data_dir,
        output_dir=args.output_dir,
        fmt=args.format,
        districts=args.districts,
        by_village=args.by_village,
        single_file=args.single_file,
    )


if __name__ == "__main__":
    main()
