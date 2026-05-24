#!/usr/bin/env python3
"""
Export completed RoR data to folder structure with Excel files.

Structure: output_dir / district / tahasil / village.xlsx

Each village Excel file contains:
- Sheet 1 "Summary": Khatiyan-level summary (one row per khatiyan)
- Sheet 2 "Plots": All plots across all khatiyans (detailed data)

Usage:
    # Export all scraped data
    python export_to_folders.py --data-dir bhulekh_data --output-dir export
    
    # Export only completed villages (checks work_queue.db)
    python export_to_folders.py --data-dir bhulekh_data --output-dir export --queue-db work_queue.db --completed-only
    
    # Export specific district
    python export_to_folders.py --data-dir bhulekh_data --output-dir export --district କଟକ
"""

import argparse
import json
import os
import sqlite3
import sys
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, Alignment, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)

import re

# Regex to match illegal XML characters that Excel can't handle
ILLEGAL_CHARS_RE = re.compile(
    r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]'
)


def clean_for_excel(value: Any) -> Any:
    """Remove illegal characters that Excel/openpyxl can't handle."""
    if isinstance(value, str):
        # Remove illegal XML characters
        value = ILLEGAL_CHARS_RE.sub('', value)
        # Also remove any null bytes
        value = value.replace('\x00', '')
        # Truncate very long strings (Excel cell limit is 32767 chars)
        if len(value) > 32000:
            value = value[:32000] + "..."
    return value


def sanitize_filename(name: str) -> str:
    """Make a string safe for use as filename."""
    # Remove/replace invalid characters
    invalid_chars = '<>:"/\\|?*'
    for char in invalid_chars:
        name = name.replace(char, '_')
    return name.strip() or "unknown"


def get_completed_villages(queue_db: str) -> Set[Tuple[str, str, str]]:
    """
    Get set of (district_name, tahasil_name, village_name) tuples 
    for villages marked as 'done' in the work queue.
    """
    completed = set()
    try:
        conn = sqlite3.connect(queue_db)
        cursor = conn.execute(
            "SELECT district_name, tahasil_name, village_name FROM villages WHERE status = 'done'"
        )
        for row in cursor:
            completed.add((row[0], row[1], row[2]))
        conn.close()
        print(f"Found {len(completed)} completed villages in work queue")
    except Exception as e:
        print(f"Error reading work queue: {e}")
    return completed


def read_sqlite_data(db_path: Path, completed_villages: Optional[Set[Tuple[str, str, str]]] = None) -> List[Dict[str, Any]]:
    """Read khatiyan data from a SQLite database.
    
    If completed_villages is provided, only returns khatiyans from those villages.
    """
    records = []
    skipped = 0
    try:
        conn = sqlite3.connect(str(db_path))
        cursor = conn.execute(
            "SELECT district, tahasil, village, khatiyan_value, khatiyan_text, data_json FROM khatiyans"
        )
        for row in cursor:
            district, tahasil, village, kh_value, kh_text, data_json = row
            
            # Filter by completed villages if specified
            if completed_villages is not None:
                if (district, tahasil, village) not in completed_villages:
                    skipped += 1
                    continue
            
            try:
                data = json.loads(data_json)
                data['district'] = district
                data['tahasil'] = tahasil
                data['village'] = village
                data['khatiyan_value'] = kh_value
                data['khatiyan_text'] = kh_text
                records.append(data)
            except json.JSONDecodeError:
                pass
        conn.close()
        if skipped > 0:
            print(f"  Skipped {skipped} khatiyans from incomplete villages")
    except Exception as e:
        print(f"Error reading {db_path}: {e}")
    return records


def read_ndjson_data(ndjson_path: Path) -> List[Dict[str, Any]]:
    """Read all khatiyan data from an NDJSON file."""
    records = []
    try:
        with open(ndjson_path, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line:
                    try:
                        data = json.loads(line)
                        records.append(data)
                    except json.JSONDecodeError:
                        pass
    except Exception as e:
        print(f"Error reading {ndjson_path}: {e}")
    return records


def group_by_village(records: List[Dict[str, Any]]) -> Dict[str, Dict[str, Dict[str, List[Dict]]]]:
    """Group records by district > tahasil > village."""
    grouped = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
    for record in records:
        district = record.get('district', 'Unknown')
        tahasil = record.get('tahasil', 'Unknown')
        village = record.get('village', 'Unknown')
        grouped[district][tahasil][village].append(record)
    return grouped


EXCEL_COLUMNS = [
    'district',
    'mouja',
    'tehsil',
    'thana',
    'tehsil_no',
    'thana_no',
    'landlord_name',
    'khatiyan_sl_no',
    'tenant_name',
    'status',
    'water_tax',
    'tax',
    'ses',
    'other_ses',
    'total',
    'description',
    'special_case',
    'last_publish_date',
    'tax_date',
    'form_no',
    'parichheda',
    'plot_plot_no',
    'plot_chaka',
    'plot_land_type',
    'plot_kisam',
    'plot_n_occu',
    'plot_e_occu',
    'plot_s_occu',
    'plot_w_occu',
    'plot_acre',
    'plot_decimil',
    'plot_hector',
    'plot_remarks',
]


def _khatiyan_base_row(kh: Dict[str, Any]) -> Dict[str, Any]:
    return {
        'district': clean_for_excel(kh.get('district', '')),
        'mouja': clean_for_excel(kh.get('mouja', '')),
        'tehsil': clean_for_excel(kh.get('tahasil', '') or kh.get('tehsil', '')),
        'thana': clean_for_excel(kh.get('thana', '')),
        'tehsil_no': clean_for_excel(kh.get('tehsil_no', '')),
        'thana_no': clean_for_excel(kh.get('thana_no', '')),
        'landlord_name': clean_for_excel(kh.get('landlord_name', '')),
        'khatiyan_sl_no': clean_for_excel(kh.get('khatiyan_sl_no', '') or kh.get('khatiyan_text', '')),
        'tenant_name': clean_for_excel(kh.get('tenant_name', '')),
        'status': clean_for_excel(kh.get('status', '')),
        'water_tax': clean_for_excel(kh.get('water_tax', '')),
        'tax': clean_for_excel(kh.get('tax', '')),
        'ses': clean_for_excel(kh.get('ses', '')),
        'other_ses': clean_for_excel(kh.get('other_ses', '')),
        'total': clean_for_excel(kh.get('total', '')),
        'description': clean_for_excel(kh.get('description', '')),
        'special_case': clean_for_excel(kh.get('special_case', '')),
        'last_publish_date': clean_for_excel(kh.get('last_publish_date', '')),
        'tax_date': clean_for_excel(kh.get('tax_date', '')),
        'form_no': clean_for_excel(kh.get('form_no', '')),
        'parichheda': clean_for_excel(kh.get('parichheda', '')),
    }


def _plot_row(base_data: Dict[str, Any], plot: Optional[Dict[str, Any]] = None) -> List[Any]:
    plot = plot or {}
    return [
        base_data['district'],
        base_data['mouja'],
        base_data['tehsil'],
        base_data['thana'],
        base_data['tehsil_no'],
        base_data['thana_no'],
        base_data['landlord_name'],
        base_data['khatiyan_sl_no'],
        base_data['tenant_name'],
        base_data['status'],
        base_data['water_tax'],
        base_data['tax'],
        base_data['ses'],
        base_data['other_ses'],
        base_data['total'],
        base_data['description'],
        base_data['special_case'],
        base_data['last_publish_date'],
        base_data['tax_date'],
        base_data['form_no'],
        base_data['parichheda'],
        clean_for_excel(plot.get('plot_no', '')),
        clean_for_excel(plot.get('chaka', '')),
        clean_for_excel(plot.get('land_type', '')),
        clean_for_excel(plot.get('kisam', '')),
        clean_for_excel(plot.get('n_occu', '')),
        clean_for_excel(plot.get('e_occu', '')),
        clean_for_excel(plot.get('s_occu', '')),
        clean_for_excel(plot.get('w_occu', '')),
        clean_for_excel(plot.get('acre', '')),
        clean_for_excel(plot.get('decimil', '')),
        clean_for_excel(plot.get('hector', '')),
        clean_for_excel(plot.get('remarks', '')),
    ]


def create_village_excel(
    village_name: str,
    khatiyans: List[Dict[str, Any]],
    output_path: Path
) -> int:
    """Create an Excel file for a village with all its khatiyans.
    
    Creates a FLAT format: one row per plot with all khatiyan metadata repeated.
    Returns the number of plots written.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    headers = EXCEL_COLUMNS
    ws.append(headers)
    
    # Style header row
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    total_plots = 0
    for kh in khatiyans:
        plots = kh.get('plots', [])
        base_data = _khatiyan_base_row(kh)

        if plots:
            for plot in plots:
                ws.append(_plot_row(base_data, plot))
                total_plots += 1
        else:
            ws.append(_plot_row(base_data))
            total_plots += 1
    
    # Auto-width columns
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Save (overwrites existing)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    wb.close()
    
    return total_plots


def export_data(
    data_dir: str,
    output_dir: str,
    filter_district: Optional[str] = None,
    filter_tahasil: Optional[str] = None,
    queue_db: Optional[str] = None,
    completed_only: bool = False,
):
    """Export data to folder structure with Excel files.
    
    If completed_only=True, only exports villages marked as 'done' in the work queue.
    Files are overwritten if they already exist.
    """
    data_path = Path(data_dir)
    output_path = Path(output_dir)
    
    if not data_path.exists():
        print(f"ERROR: Data directory not found: {data_dir}")
        return
    
    # Get completed villages if filtering
    completed_villages = None
    if completed_only:
        if not queue_db:
            print("ERROR: --queue-db required when using --completed-only")
            return
        if not Path(queue_db).exists():
            print(f"ERROR: Work queue not found: {queue_db}")
            return
        completed_villages = get_completed_villages(queue_db)
        if not completed_villages:
            print("No completed villages found in work queue")
            return
    
    # Find all data files
    db_files = list(data_path.glob("district_*.db"))
    ndjson_files = list(data_path.glob("district_*.ndjson"))
    
    if not db_files and not ndjson_files:
        print(f"No data files found in {data_dir}")
        return
    
    mode = "COMPLETED ONLY" if completed_only else "ALL DATA"
    print(f"Export mode: {mode}")
    print(f"Found {len(db_files)} SQLite files and {len(ndjson_files)} NDJSON files")
    
    # Read all data
    all_records = []
    for db_file in db_files:
        print(f"Reading {db_file.name}...")
        records = read_sqlite_data(db_file, completed_villages)
        all_records.extend(records)
        print(f"  -> {len(records)} khatiyans")
    
    for ndjson_file in ndjson_files:
        print(f"Reading {ndjson_file.name}...")
        records = read_ndjson_data(ndjson_file)
        # Filter NDJSON records too if needed
        if completed_villages is not None:
            records = [r for r in records 
                      if (r.get('district'), r.get('tahasil'), r.get('village')) in completed_villages]
        all_records.extend(records)
        print(f"  -> {len(records)} khatiyans")
    
    print(f"\nTotal khatiyans loaded: {len(all_records)}")
    
    # Group by district > tahasil > village
    grouped = group_by_village(all_records)
    
    # Apply filters
    if filter_district:
        if filter_district in grouped:
            grouped = {filter_district: grouped[filter_district]}
        else:
            print(f"District '{filter_district}' not found in data")
            return
    
    # Export
    total_villages = 0
    total_khatiyans = 0
    total_plots = 0
    
    for district, tahasils in sorted(grouped.items()):
        district_safe = sanitize_filename(district)
        
        for tahasil, villages in sorted(tahasils.items()):
            if filter_tahasil and tahasil != filter_tahasil:
                continue
                
            tahasil_safe = sanitize_filename(tahasil)
            
            for village, khatiyans in sorted(villages.items()):
                village_safe = sanitize_filename(village)
                
                # Create Excel file path
                excel_path = output_path / district_safe / tahasil_safe / f"{village_safe}.xlsx"
                
                # Create the Excel file
                plots_count = create_village_excel(village, khatiyans, excel_path)
                
                total_villages += 1
                total_khatiyans += len(khatiyans)
                total_plots += plots_count
                
                print(f"  {district}/{tahasil}/{village}: {len(khatiyans)} khatiyans, {plots_count} plots")
    
    print(f"\n{'='*60}")
    print(f"EXPORT COMPLETE")
    print(f"{'='*60}")
    print(f"Output directory: {output_path}")
    print(f"Villages exported: {total_villages}")
    print(f"Total khatiyans: {total_khatiyans}")
    print(f"Total plots: {total_plots}")


def main():
    parser = argparse.ArgumentParser(
        description="Export RoR data to folder structure with Excel files"
    )
    parser.add_argument(
        "--data-dir", default="bhulekh_data",
        help="Directory containing district SQLite/NDJSON files (default: bhulekh_data)"
    )
    parser.add_argument(
        "--output-dir", default="export",
        help="Output directory for Excel files (default: export). Existing files are overwritten."
    )
    parser.add_argument(
        "--district", default=None,
        help="Filter to specific district name (Odia text)"
    )
    parser.add_argument(
        "--tahasil", default=None,
        help="Filter to specific tahasil name (Odia text)"
    )
    parser.add_argument(
        "--queue-db", default=None,
        help="Path to work queue database (required for --completed-only)"
    )
    parser.add_argument(
        "--completed-only", action="store_true",
        help="Only export villages marked as 'done' in the work queue"
    )
    
    args = parser.parse_args()
    
    export_data(
        data_dir=args.data_dir,
        output_dir=args.output_dir,
        filter_district=args.district,
        filter_tahasil=args.tahasil,
        queue_db=args.queue_db,
        completed_only=args.completed_only,
    )


if __name__ == "__main__":
    main()
